"""Personal finance — snap a bill, Himmy reads & files it; track spending; Excel in & out.

The magic moment: you photograph a receipt, Himmy reads it (the image OCR we already have),
turns it into a structured expense (the framework's own inference — so cost is metered into
/usage like everything else), and files it in a ledger it owns. From there it can tell you where
your money went, and you can import an existing spreadsheet or export the ledger to Excel/CSV so
you're never locked in.

Reuses, never rebuilds:
  * :func:`himmy_app.connectors.media.image_to_text` — the bill photo → text;
  * the (extended) framework reader factory — PDF / CSV / XLSX statements → text;
  * ``himmy.cli.provider.build_inference_for`` — the receipt text → a clean JSON expense.

The ledger (``finance.db``) is a normal store under the workspace, so it rides the backup.
"""

from __future__ import annotations

import contextlib
import csv
import datetime
import io
import json
import os
import re
import sqlite3
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any

from himmy_app.config import HimmyConfig, load_config

#: The fixed category set — small enough to stay consistent, broad enough to cover daily life.
CATEGORIES = [
    "Food", "Groceries", "Transport", "Shopping", "Bills", "Health",
    "Entertainment", "Travel", "Education", "Other",
]
_CATSET = {c.lower(): c for c in CATEGORIES}
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _today() -> str:
    return datetime.date.today().isoformat()


def _to_amount(v: Any) -> float:
    """Parse a money value out of a number or a string like 'Rs 1,250.50' → 1250.5."""
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    m = re.search(r"(\d[\d,]*\.?\d*)", str(v or ""))
    if not m:
        return 0.0
    try:
        return round(float(m.group(1).replace(",", "")), 2)
    except (TypeError, ValueError):
        return 0.0


def _clean_category(v: Any) -> str:
    return _CATSET.get(str(v or "").strip().lower(), "Other")


def _clean_date(v: Any) -> str:
    s = str(v or "").strip()[:10]
    return s if _DATE_RE.match(s) else _today()


class ExpenseStore:
    """The expense ledger (``finance.db``) — Himmy-owned, backup-covered."""

    def __init__(self, config: HimmyConfig | None = None) -> None:
        cfg = config or load_config()
        self._cfg = cfg
        self._db = cfg.data_dir / "finance.db"
        self._ensure()

    def _conn(self) -> sqlite3.Connection:
        c = sqlite3.connect(str(self._db), timeout=10)
        c.row_factory = sqlite3.Row
        return c

    def _ensure(self) -> None:
        with self._conn() as c:
            c.execute(
                """CREATE TABLE IF NOT EXISTS expenses (
                    id TEXT PRIMARY KEY, date TEXT, merchant TEXT, amount REAL,
                    currency TEXT, category TEXT, note TEXT, items TEXT,
                    source TEXT, created REAL
                )"""
            )

    @staticmethod
    def _new_id() -> str:
        # Time-prefixed (so ids sort roughly by creation) + a random suffix so two expenses added in
        # the SAME millisecond (a quick add, or a CSV/Excel import loop) never collide and silently
        # overwrite each other via INSERT OR REPLACE.
        return f"exp_{int(time.time() * 1000):x}_{uuid.uuid4().hex[:8]}"

    # ---- writes -------------------------------------------------------------------------
    def add(self, e: dict[str, Any], *, source: str = "manual") -> dict[str, Any]:
        row = {
            "id": self._new_id(),
            "date": _clean_date(e.get("date")),
            "merchant": str(e.get("merchant") or "").strip()[:120] or "Expense",
            "amount": _to_amount(e.get("amount")),
            "currency": (str(e.get("currency") or "NPR").strip()[:8] or "NPR").upper(),
            "category": _clean_category(e.get("category")),
            "note": str(e.get("note") or "").strip()[:300],
            "items": json.dumps([str(x).strip()[:80] for x in (e.get("items") or []) if str(x).strip()][:12]),
            "source": source,
            "created": time.time(),
        }
        with self._conn() as c:
            c.execute(
                """INSERT OR REPLACE INTO expenses
                   (id,date,merchant,amount,currency,category,note,items,source,created)
                   VALUES (:id,:date,:merchant,:amount,:currency,:category,:note,:items,:source,:created)""",
                row,
            )
        return self._row(row)

    def add_many(self, rows: list[dict[str, Any]], *, source: str = "excel") -> int:
        n = 0
        for r in rows:
            if _to_amount(r.get("amount")) > 0:
                self.add(r, source=source)
                n += 1
        return n

    def delete(self, exp_id: str) -> dict[str, Any]:
        with self._conn() as c:
            c.execute("DELETE FROM expenses WHERE id = ?", (exp_id,))
        return {"ok": True}

    def clear(self, *, source: str | None = None) -> int:
        """Delete all expenses (or only those from one ``source``). Returns how many were removed."""
        with self._conn() as c:
            if source:
                cur = c.execute("DELETE FROM expenses WHERE source = ?", (source,))
            else:
                cur = c.execute("DELETE FROM expenses")
            return int(cur.rowcount or 0)

    # ---- reads --------------------------------------------------------------------------
    def _row(self, r: Any) -> dict[str, Any]:
        d = dict(r)
        try:
            d["items"] = json.loads(d.get("items") or "[]")
        except Exception:  # noqa: BLE001
            d["items"] = []
        return d

    def list(self, *, limit: int = 300, month: str | None = None,
             category: str | None = None) -> list[dict[str, Any]]:
        q = "SELECT * FROM expenses"
        clauses, params = [], []
        if month:
            clauses.append("date LIKE ?")
            params.append(f"{month}%")
        if category:
            clauses.append("category = ?")
            params.append(_clean_category(category))
        if clauses:
            q += " WHERE " + " AND ".join(clauses)
        q += " ORDER BY date DESC, created DESC LIMIT ?"
        params.append(int(limit))
        with self._conn() as c:
            return [self._row(r) for r in c.execute(q, params).fetchall()]

    def months(self) -> list[str]:
        with self._conn() as c:
            return [r[0] for r in c.execute(
                "SELECT DISTINCT substr(date,1,7) m FROM expenses ORDER BY m DESC").fetchall() if r[0]]

    def _range_for(self, period: str) -> tuple[str | None, str | None, str]:
        """(from_date, to_date, label) for a period keyword. None bounds = open-ended."""
        today = datetime.date.today()
        p = (period or "month").lower().replace("this_", "").replace("this ", "").strip()
        if p in ("week", "7d", "7days"):
            start = today - datetime.timedelta(days=6)
            return start.isoformat(), today.isoformat(), "the last 7 days"
        if p in ("year", "ytd"):
            return f"{today.year}-01-01", today.isoformat(), f"{today.year}"
        if p in ("all", "alltime", "everything"):
            return None, None, "all time"
        # default: this calendar month
        return today.replace(day=1).isoformat(), today.isoformat(), today.strftime("%B %Y")

    def summary(self, period: str = "month") -> dict[str, Any]:
        frm, to, label = self._range_for(period)
        q = "SELECT category, COUNT(*) n, COALESCE(SUM(amount),0) t, currency FROM expenses"
        clauses, params = [], []
        if frm:
            clauses.append("date >= ?")
            params.append(frm)
        if to:
            clauses.append("date <= ?")
            params.append(to)
        if clauses:
            q += " WHERE " + " AND ".join(clauses)
        q += " GROUP BY category ORDER BY t DESC"
        by_cat: list[dict[str, Any]] = []
        total = 0.0
        count = 0
        currency = "NPR"
        with self._conn() as c:
            for r in c.execute(q, params).fetchall():
                by_cat.append({"category": r["category"], "total": round(r["t"], 2), "count": r["n"]})
                total += r["t"]
                count += r["n"]
                if r["currency"]:
                    currency = r["currency"]
        return {"ok": True, "period": period, "label": label, "from": frm, "to": to,
                "total": round(total, 2), "count": count, "currency": currency, "by_category": by_cat}

    # ---- Excel / CSV in & out -----------------------------------------------------------
    def export(self, *, fmt: str = "xlsx") -> dict[str, Any]:
        """Write the whole ledger to ~/Downloads as .xlsx (openpyxl) or .csv; return the path."""
        rows = self.list(limit=100000)
        dest = Path("~/Downloads").expanduser()
        dest.mkdir(parents=True, exist_ok=True)
        stamp = time.strftime("%Y%m%d-%H%M%S")
        headers = ["date", "merchant", "amount", "currency", "category", "note", "source"]
        if fmt == "xlsx":
            try:
                import openpyxl

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Expenses"
                ws.append([h.capitalize() for h in headers])
                for r in rows:
                    ws.append([r.get(h) for h in headers])
                path = dest / f"himmy-finance-{stamp}.xlsx"
                wb.save(str(path))
                return {"ok": True, "path": str(path)}
            except Exception:  # noqa: BLE001 - fall back to CSV if openpyxl misbehaves
                pass
        path = dest / f"himmy-finance-{stamp}.csv"
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h) for h in headers})
        return {"ok": True, "path": str(path)}

    def import_bytes(self, data: bytes, name: str = "") -> dict[str, Any]:
        """Import expenses from an uploaded CSV or XLSX (flexible column names). Returns count."""
        ext = Path(name or "").suffix.lower()
        rows: list[dict[str, str]] = []
        try:
            if ext in (".xlsx", ".xlsm"):
                import openpyxl

                wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
                ws = wb.active
                it = ws.iter_rows(values_only=True)
                header = [str(h or "").strip().lower() for h in next(it, [])]
                for raw in it:
                    rows.append({header[i]: ("" if v is None else str(v))
                                 for i, v in enumerate(raw) if i < len(header)})
            else:  # CSV (default)
                text = data.decode("utf-8", "ignore")
                for r in csv.DictReader(io.StringIO(text)):
                    rows.append({str(k or "").strip().lower(): str(v or "") for k, v in r.items()})
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "message": f"Couldn't read that file ({type(exc).__name__})."}

        def pick(d: dict[str, str], *names: str) -> str:
            for n in names:
                for k, v in d.items():
                    if n in k and v:
                        return v
            return ""

        mapped = [{
            "date": pick(r, "date"),
            "merchant": pick(r, "merchant", "description", "name", "details", "payee", "item"),
            "amount": pick(r, "amount", "total", "price", "debit", "spent", "cost"),
            "category": pick(r, "category", "type", "tag"),
            "note": pick(r, "note", "remarks", "memo"),
        } for r in rows]
        n = self.add_many([m for m in mapped if _to_amount(m["amount"]) > 0], source="excel")
        return {"ok": True, "imported": n, "rows": len(rows)}


# --------------------------------------------------------------------------------------------
# Snap-a-bill: image/PDF bytes → text → a structured expense draft (framework inference)
# --------------------------------------------------------------------------------------------
_BILL_SYS = (
    "You extract ONE expense from the text of a receipt or bill. Return ONLY a JSON object with "
    'these keys: {"merchant": string, "date": "YYYY-MM-DD" or "", "amount": number, "currency": '
    'string, "category": one of ' + json.dumps(CATEGORIES) + ', "items": [up to 6 short strings], '
    '"note": string}. RULES: "amount" is the final grand TOTAL actually paid (including tax/service '
    "charge), NOT a single line item. Infer the most fitting category from the merchant and items. "
    "Default currency to NPR if a Nepali/Rs bill. If the date isn't on the bill use \"\". Output the "
    "JSON object and nothing else."
)


async def _bill_text(data: bytes, mime: str, name: str, cfg: HimmyConfig) -> str:
    """Get the readable text of a bill: OCR for an image, the framework reader for a doc/PDF."""
    m = (mime or "").lower()
    img_exts = {".png", ".jpg", ".jpeg", ".heic", ".heif", ".webp", ".gif", ".tif", ".tiff", ".bmp"}
    if m.startswith("image/") or Path(name or "").suffix.lower() in img_exts:
        from himmy_app.connectors.media import image_to_text

        return await image_to_text(data, mime or "image/jpeg", cfg)
    ext = Path(name or "").suffix.lower()
    if ext in {".pdf", ".txt", ".md", ".csv", ".xlsx", ".xlsm", ".docx", ".html", ".htm"}:
        tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
        try:
            tmp.write(data)
            tmp.close()
            from himmy_app.attachments import _factory

            return _factory().read(tmp.name)
        except Exception:  # noqa: BLE001
            pass
        finally:
            with contextlib.suppress(Exception):
                os.unlink(tmp.name)
    return data.decode("utf-8", "ignore")


def _parse_expense_json(content: str) -> dict[str, Any] | None:
    m = re.search(r"\{.*\}", content or "", re.DOTALL)
    if not m:
        return None
    try:
        d = json.loads(m.group(0))
    except Exception:  # noqa: BLE001
        return None
    return d if isinstance(d, dict) else None


async def extract_bill(data: bytes, mime: str = "", name: str = "",
                       cfg: HimmyConfig | None = None) -> dict[str, Any]:
    """Read a bill (photo/PDF/text) into a structured expense DRAFT (not yet saved).

    Returns ``{ok, expense: {date,merchant,amount,currency,category,items,note}, text}``. The
    draft is for the user to confirm before it's filed. Fully best-effort.
    """
    cfg = cfg or load_config()
    text = (await _bill_text(data, mime, name, cfg) or "").strip()
    if not text:
        return {"ok": False, "message": "Couldn't read that bill — the current model may not read "
                "images. Switch to OpenRouter (gemini-2.5-flash) in Settings → Preferences."}
    try:
        from himmy.cli.provider import build_inference_for
        from himmy.services.inference.models import InferenceMessage, InferenceRequest

        svc = build_inference_for(cfg.provider, cfg.model)
        resp = await svc.run(InferenceRequest(
            messages=[InferenceMessage(role="system", content=_BILL_SYS),
                      InferenceMessage(role="user", content=text[:6000])],
            generation_params={"temperature": 0}, timeout_seconds=45,
        ))
        parsed = _parse_expense_json(resp.output_text or "")
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "message": f"Couldn't understand the bill ({type(exc).__name__}).",
                "text": text[:600]}
    if not parsed:
        return {"ok": False, "message": "Couldn't pull an amount off that bill — add it manually?",
                "text": text[:600]}
    expense = {
        "date": _clean_date(parsed.get("date")),
        "merchant": str(parsed.get("merchant") or "").strip()[:120] or "Expense",
        "amount": _to_amount(parsed.get("amount")),
        "currency": (str(parsed.get("currency") or "NPR").strip()[:8] or "NPR").upper(),
        "category": _clean_category(parsed.get("category")),
        "items": [str(x).strip()[:80] for x in (parsed.get("items") or []) if str(x).strip()][:8],
        "note": str(parsed.get("note") or "").strip()[:300],
    }
    return {"ok": True, "expense": expense, "text": text[:600]}


__all__ = ["ExpenseStore", "extract_bill", "CATEGORIES"]
